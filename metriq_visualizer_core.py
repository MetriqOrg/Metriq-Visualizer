# Copyright (c) Metriq Foundation, Inc.
# This Source Code Form is subject to the terms of the Mozilla Public License, v. 2.0.
"""Local media/table analysis, formula evaluation, and geometry construction.

This module deliberately has no Qt dependency.  It is the stable computational
boundary used by the desktop shell, renderer, export worker, tests, and future
headless integrations.
"""

from __future__ import annotations

import ast
import csv
import io
import math
import re
import shutil
import subprocess
from collections.abc import Callable, Mapping, Sequence
from contextlib import suppress
from dataclasses import dataclass, field
from math import gcd
from pathlib import Path
from typing import Any

import numpy as np

TABLE_EXTENSIONS = {".csv", ".tsv", ".txt", ".xlsx"}
VIDEO_EXTENSIONS = {".mp4", ".mov", ".avi", ".mkv", ".webm", ".m4v", ".mpg", ".mpeg", ".wmv", ".flv"}
AUDIO_EXTENSIONS = {".mp3", ".wav", ".flac", ".ogg", ".m4a", ".aac", ".aiff", ".aif", ".opus"}
ANALYSIS_ENGINE_VERSION = "numpy-scipy-bounded-v3"
MAX_ANALYSIS_FRAMES = 4096

DEFAULT_PRESETS: dict[str, dict[str, str]] = {
    # Original Metriq mappings are intentionally retained.  Existing projects
    # and old .mvpreset files depend on these names and formula aliases.
    "Pitch/Timbre/Motion": {
        "x": "0.7*chroma_mean + 0.3*f0_hz",
        "y": "0.6*mfcc_1 + 0.4*mfcc_2",
        "z": "0.6*spectral_flux + 0.4*onset_strength",
        "color": "spectral_centroid_hz",
        "size": "rms",
    },
    "Birdsong": {
        "x": "0.65*f0_hz + 0.35*chroma_mean",
        "y": "0.55*mfcc_1 + 0.30*mfcc_2 + 0.15*mfcc_3",
        "z": "0.55*spectral_flux + 0.30*onset_strength + 0.15*zcr",
        "color": "spectral_centroid_hz",
        "size": "0.55*rms + 0.45*onset_strength",
    },
    "Audio PCA": {
        "x": "pc1",
        "y": "pc2",
        "z": "pc3",
        "color": "f0_hz",
        "size": "0.65*rms + 0.35*onset_strength",
    },
    "Rhythm / Brightness / Texture": {
        "x": "0.7*onset_strength + 0.3*spectral_flux",
        "y": "spectral_centroid_hz",
        "z": "0.6*spectral_flatness + 0.4*chroma_entropy",
        "color": "dominant_freq_hz",
        "size": "0.7*rms + 0.3*zcr",
    },
    "Spectral orbit": {
        "x": "spectral_centroid",
        "y": "spectral_bandwidth",
        "z": "spectral_flux",
        "color": "dominant_frequency",
        "size": "rms",
    },
    "Timbre lattice": {
        "x": "mfcc_1",
        "y": "mfcc_2",
        "z": "mfcc_3",
        "color": "chroma_mean",
        "size": "rms",
    },
    "Rhythm trail": {
        "x": "time",
        "y": "onset_strength",
        "z": "spectral_flux",
        "color": "rms_db",
        "size": "onset_strength",
    },
    "Table / PCA explorer": {
        "x": "pc1",
        "y": "pc2",
        "z": "pc3",
        "color": "time",
        "size": "magnitude",
    },
    "Table / Spread / Change": {
        "x": "column_mean",
        "y": "column_spread",
        "z": "delta_magnitude",
        "color": "time",
        "size": "magnitude",
    },
    "Manual starter": {
        "x": "pc1 + 0.2*mfcc_1",
        "y": "pc2 + 0.3*chroma_mean",
        "z": "pc3 + 0.2*spectral_flux",
        "color": "dominant_freq_hz",
        "size": "0.5*rms + 0.5*onset_strength",
    },
}

FEATURE_DESCRIPTIONS: dict[str, str] = {
    "time": "Frame timestamp in seconds.",
    "frame": "Zero-based analysis-frame index.",
    "rms": "Root-mean-square signal energy.",
    "rms_db": "Signal energy in decibels relative to the source peak.",
    "zero_crossing_rate": "Rate at which the waveform changes sign.",
    "spectral_centroid": "Energy-weighted center frequency in hertz.",
    "spectral_bandwidth": "Spread of spectral energy around the centroid.",
    "spectral_rolloff": "Frequency below which 85% of spectral energy lies.",
    "spectral_flatness": "Noise-like versus tonal spectral character.",
    "spectral_flux": "Frame-to-frame spectral change.",
    "onset_strength": "Local transient/onset intensity.",
    "dominant_frequency": "Strongest FFT-bin frequency in hertz.",
    "fundamental_frequency": "Bounded harmonic-summation estimate of the fundamental frequency in hertz.",
    "chroma_mean": "Average energy across the twelve pitch classes.",
    "pc1": "First principal component of the available numeric features.",
    "pc2": "Second principal component of the available numeric features.",
    "pc3": "Third principal component of the available numeric features.",
    "magnitude": "Euclidean magnitude across imported numeric columns.",
    "column_mean": "Mean across imported numeric columns for each row.",
    "column_spread": "Standard deviation across imported numeric columns for each row.",
    "delta_magnitude": "Absolute row-to-row change in magnitude.",
    "t": "Alias for frame timestamp in seconds.",
    "row_index": "Alias for the analysis-frame or imported-row index.",
    "zcr": "Legacy alias for zero-crossing rate.",
    "spectral_centroid_hz": "Legacy hertz alias for spectral centroid.",
    "spectral_bandwidth_hz": "Legacy hertz alias for spectral bandwidth.",
    "spectral_rolloff_hz": "Legacy hertz alias for spectral rolloff.",
    "dominant_freq_hz": "Legacy hertz alias for dominant frequency.",
    "f0_hz": "Legacy alias for the bounded fundamental-frequency estimate.",
    "chroma_entropy": "Normalized entropy across the twelve pitch classes.",
    "spectral_contrast_mean": "Fast contrast proxy derived from spectral flatness.",
}


_LEGACY_FEATURE_ALIASES: dict[str, str] = {
    "t": "time",
    "row_index": "frame",
    "zcr": "zero_crossing_rate",
    "spectral_centroid_hz": "spectral_centroid",
    "spectral_bandwidth_hz": "spectral_bandwidth",
    "spectral_rolloff_hz": "spectral_rolloff",
    "dominant_freq_hz": "dominant_frequency",
    "f0_hz": "fundamental_frequency",
}


def _install_legacy_feature_aliases(features: dict[str, np.ndarray], length: int) -> None:
    """Install historical formula names without duplicating analysis work."""

    if "fundamental_frequency" not in features and "dominant_frequency" in features:
        features["fundamental_frequency"] = _align_vector(features["dominant_frequency"], length)

    for alias, source in _LEGACY_FEATURE_ALIASES.items():
        if alias not in features and source in features:
            features[alias] = _align_vector(features[source], length)

    if "chroma_entropy" not in features:
        channels = [features.get(f"chroma_{index}") for index in range(1, 13)]
        if all(channel is not None for channel in channels):
            chroma = np.vstack([_align_vector(channel, length) for channel in channels if channel is not None])
            chroma = np.clip(chroma, 0.0, None)
            chroma /= np.maximum(np.sum(chroma, axis=0, keepdims=True), 1e-20)
            entropy = -np.sum(chroma * np.log(np.maximum(chroma, 1e-20)), axis=0) / np.log(12.0)
            features["chroma_entropy"] = entropy.astype(np.float32)
        else:
            features["chroma_entropy"] = np.zeros(length, dtype=np.float32)

    if "spectral_contrast_mean" not in features:
        flatness = _align_vector(features.get("spectral_flatness", np.zeros(length)), length)
        finite = np.clip(flatness, 0.0, 1.0)
        features["spectral_contrast_mean"] = (1.0 - finite).astype(np.float32)


def _safe_setting_int(value: Any, default: int) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return int(default)


def _safe_setting_float(value: Any, default: float) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return float(default)


@dataclass(frozen=True, slots=True)
class AnalysisSettings:
    """Creator-controlled extraction settings with conservative local bounds."""

    sample_rate: int = 0
    n_fft: int = 2048
    hop_length: int = 512
    min_frequency: float = 0.0
    max_frequency: float = 0.0
    max_frames: int = MAX_ANALYSIS_FRAMES
    n_mels: int = 96
    mfcc_count: int = 20

    def normalized(self, *, native_sample_rate: int | None = None) -> AnalysisSettings:
        sample_rate = int(self.sample_rate)
        if sample_rate != 0:
            sample_rate = int(np.clip(sample_rate, 8_000, 192_000))
        elif native_sample_rate is not None:
            sample_rate = 0
        n_fft = int(np.clip(self.n_fft, 256, 16_384))
        n_fft = 1 << int(round(math.log2(max(256, n_fft))))
        n_fft = int(np.clip(n_fft, 256, 16_384))
        hop = int(np.clip(self.hop_length, 16, n_fft))
        minimum = max(0.0, float(self.min_frequency))
        nyquist = (sample_rate or int(native_sample_rate or 192_000)) / 2.0
        maximum = float(self.max_frequency)
        maximum = (
            0.0
            if maximum <= 0.0
            else float(np.clip(maximum, minimum + 1.0, max(minimum + 1.0, nyquist)))
        )
        return AnalysisSettings(
            sample_rate=sample_rate,
            n_fft=n_fft,
            hop_length=hop,
            min_frequency=minimum,
            max_frequency=maximum,
            max_frames=int(np.clip(self.max_frames, 128, 16_384)),
            n_mels=int(np.clip(self.n_mels, 24, 256)),
            mfcc_count=int(np.clip(self.mfcc_count, 4, 40)),
        )

    def to_dict(self) -> dict[str, int | float]:
        return {
            "sample_rate": int(self.sample_rate),
            "n_fft": int(self.n_fft),
            "hop_length": int(self.hop_length),
            "min_frequency": float(self.min_frequency),
            "max_frequency": float(self.max_frequency),
            "max_frames": int(self.max_frames),
            "n_mels": int(self.n_mels),
            "mfcc_count": int(self.mfcc_count),
        }

    @classmethod
    def from_mapping(cls, values: Mapping[str, Any] | None) -> AnalysisSettings:
        data = values or {}
        return cls(
            sample_rate=_safe_setting_int(data.get("sample_rate", data.get("sample_rate_hz")), 0),
            n_fft=_safe_setting_int(data.get("n_fft", data.get("fft_size")), 2048),
            hop_length=_safe_setting_int(data.get("hop_length", data.get("hop")), 512),
            min_frequency=_safe_setting_float(data.get("min_frequency", data.get("fmin")), 0.0),
            max_frequency=_safe_setting_float(data.get("max_frequency", data.get("fmax")), 0.0),
            max_frames=_safe_setting_int(
                data.get("max_frames", data.get("max_analysis_frames")), MAX_ANALYSIS_FRAMES
            ),
            n_mels=_safe_setting_int(data.get("n_mels"), 96),
            mfcc_count=_safe_setting_int(data.get("mfcc_count", data.get("n_mfcc")), 20),
        ).normalized()

    def signature(self) -> str:
        values = self.normalized().to_dict()
        return ";".join(f"{key}={values[key]}" for key in sorted(values))


@dataclass(slots=True)
class AnalysisResult:
    """Time-aligned source analysis consumed by mapping and rendering."""

    source_path: Path
    source_kind: str
    times: np.ndarray
    duration: float
    features: dict[str, np.ndarray]
    sample_rate: int = 0
    audio_path: Path | None = None
    has_video: bool = False
    spectrogram: np.ndarray = field(default_factory=lambda: np.empty((0, 0), dtype=np.float32))
    spectrogram_frequencies: np.ndarray = field(default_factory=lambda: np.empty(0, dtype=np.float32))
    chromagram: np.ndarray = field(default_factory=lambda: np.empty((0, 0), dtype=np.float32))
    mfcc: np.ndarray = field(default_factory=lambda: np.empty((0, 0), dtype=np.float32))
    waveform: np.ndarray = field(default_factory=lambda: np.empty(0, dtype=np.float32))
    metadata: dict[str, Any] = field(default_factory=dict)
    feature_descriptions: dict[str, str] = field(default_factory=dict)

    def __post_init__(self) -> None:
        self.source_path = Path(self.source_path)
        self.times = _vector(self.times)
        self.duration = float(max(0.0, self.duration))
        self.features = {str(key): _align_vector(value, self.times.size) for key, value in self.features.items()}
        _install_legacy_feature_aliases(self.features, self.times.size)
        self.spectrogram = np.asarray(self.spectrogram, dtype=np.float32)
        self.spectrogram_frequencies = _vector(self.spectrogram_frequencies)
        self.chromagram = np.asarray(self.chromagram, dtype=np.float32)
        self.mfcc = np.asarray(self.mfcc, dtype=np.float32)
        self.waveform = _vector(self.waveform)
        if self.audio_path is not None:
            self.audio_path = Path(self.audio_path)


@dataclass(slots=True)
class GeometryResult:
    """Mapped and sampled visual geometry."""

    x_full: np.ndarray
    y_full: np.ndarray
    z_full: np.ndarray
    color_full: np.ndarray
    size_full: np.ndarray
    times_full: np.ndarray
    rgba_full: np.ndarray
    source_indices_full: np.ndarray
    x_plot: np.ndarray
    y_plot: np.ndarray
    z_plot: np.ndarray
    color_plot: np.ndarray
    size_plot: np.ndarray
    times_plot: np.ndarray
    rgba_plot: np.ndarray
    source_indices_plot: np.ndarray
    formulas: dict[str, str] = field(default_factory=dict)
    normalize_mode: str = "zscore"
    colormap: str = "plasma"

    def __post_init__(self) -> None:
        for name in (
            "x_full", "y_full", "z_full", "color_full", "size_full", "times_full",
            "x_plot", "y_plot", "z_plot", "color_plot", "size_plot", "times_plot",
        ):
            setattr(self, name, _vector(getattr(self, name)))
        self.rgba_full = np.asarray(self.rgba_full, dtype=np.float32)
        self.rgba_plot = np.asarray(self.rgba_plot, dtype=np.float32)
        self.source_indices_full = np.asarray(self.source_indices_full, dtype=np.int64).reshape(-1)
        self.source_indices_plot = np.asarray(self.source_indices_plot, dtype=np.int64).reshape(-1)


# ---------------------------------------------------------------------------
# General utilities


def is_table_file(path: str | Path) -> bool:
    return Path(path).suffix.lower() in TABLE_EXTENSIONS


def _vector(value: Any) -> np.ndarray:
    array = np.asarray(value, dtype=np.float32).reshape(-1)
    if array.size:
        finite = np.isfinite(array)
        if not finite.all():
            replacement = float(np.nanmedian(array[finite])) if finite.any() else 0.0
            array = np.where(finite, array, replacement).astype(np.float32, copy=False)
    return array


def _align_vector(value: Any, length: int) -> np.ndarray:
    array = _vector(value)
    if length <= 0:
        return np.empty(0, dtype=np.float32)
    if array.size == length:
        return array
    if array.size == 0:
        return np.zeros(length, dtype=np.float32)
    if array.size == 1:
        return np.full(length, float(array[0]), dtype=np.float32)
    old_x = np.linspace(0.0, 1.0, array.size, dtype=np.float64)
    new_x = np.linspace(0.0, 1.0, length, dtype=np.float64)
    return np.interp(new_x, old_x, array).astype(np.float32)


def _safe_identifier(name: str, used: set[str]) -> str:
    candidate = re.sub(r"[^0-9A-Za-z_]+", "_", str(name).strip()).strip("_").lower()
    if not candidate:
        candidate = "input"
    if candidate[0].isdigit():
        candidate = f"input_{candidate}"
    base = candidate
    index = 2
    while candidate in used:
        candidate = f"{base}_{index}"
        index += 1
    used.add(candidate)
    return candidate


def _robust_standardize(matrix: np.ndarray) -> np.ndarray:
    matrix = np.asarray(matrix, dtype=np.float64)
    if matrix.ndim == 1:
        matrix = matrix[:, None]
    med = np.nanmedian(matrix, axis=0)
    matrix = np.where(np.isfinite(matrix), matrix, med)
    mean = np.mean(matrix, axis=0)
    std = np.std(matrix, axis=0)
    std = np.where(std > 1e-12, std, 1.0)
    return (matrix - mean) / std


def _principal_components(matrix: np.ndarray, count: int = 3) -> np.ndarray:
    values = _robust_standardize(matrix)
    rows = values.shape[0]
    if rows == 0:
        return np.empty((0, count), dtype=np.float32)
    if values.shape[1] == 0:
        return np.zeros((rows, count), dtype=np.float32)
    values = values - values.mean(axis=0, keepdims=True)
    try:
        u, singular, _vt = np.linalg.svd(values, full_matrices=False)
        scores = u * singular[None, :]
    except np.linalg.LinAlgError:
        scores = values
    result = np.zeros((rows, count), dtype=np.float32)
    available = min(count, scores.shape[1])
    if available:
        result[:, :available] = scores[:, :available].astype(np.float32)
    return result


def _normalize_panel(matrix: np.ndarray) -> np.ndarray:
    values = np.asarray(matrix, dtype=np.float32)
    if values.size == 0:
        return values
    finite = np.isfinite(values)
    if not finite.any():
        return np.zeros_like(values)
    lo, hi = np.percentile(values[finite], [2.0, 98.0])
    if hi <= lo + 1e-12:
        return np.zeros_like(values)
    return np.clip((values - lo) / (hi - lo), 0.0, 1.0).astype(np.float32)


# ---------------------------------------------------------------------------
# Formula evaluator


def _smooth(values: Any, window: Any = 5) -> np.ndarray:
    array = _vector(values)
    size = max(1, int(round(float(np.asarray(window).reshape(-1)[0]))))
    if size <= 1 or array.size <= 2:
        return array
    size = min(size, array.size)
    kernel = np.ones(size, dtype=np.float64) / size
    padded = np.pad(array, (size // 2, size - 1 - size // 2), mode="edge")
    return np.convolve(padded, kernel, mode="valid").astype(np.float32)


def _aggregate(function: Callable[..., np.ndarray], *values: Any) -> np.ndarray:
    arrays = [np.asarray(value, dtype=np.float64) for value in values]
    if not arrays:
        return np.asarray(0.0)
    broadcast = np.broadcast_arrays(*arrays)
    return function(np.stack(broadcast, axis=0), axis=0)


_FORMULA_FUNCTIONS: dict[str, Callable[..., Any]] = {
    "abs": np.abs,
    "sqrt": lambda value: np.sqrt(np.clip(value, 0.0, None)),
    "log": lambda value: np.log(np.clip(value, 1e-12, None)),
    "log1p": lambda value: np.log1p(np.clip(value, -0.999999, None)),
    "exp": lambda value: np.exp(np.clip(value, -60.0, 60.0)),
    "clip": np.clip,
    "smooth": _smooth,
    "mean": lambda *values: _aggregate(np.mean, *values),
    "avg": lambda *values: _aggregate(np.mean, *values),
    "sum": lambda *values: _aggregate(np.sum, *values),
    "max": lambda *values: _aggregate(np.max, *values),
    "min": lambda *values: _aggregate(np.min, *values),
}

_ALLOWED_BINOPS: dict[type[ast.operator], Callable[[Any, Any], Any]] = {
    ast.Add: lambda a, b: a + b,
    ast.Sub: lambda a, b: a - b,
    ast.Mult: lambda a, b: a * b,
    ast.Div: lambda a, b: a / np.where(np.abs(b) > 1e-12, b, np.sign(b) * 1e-12 + (b == 0) * 1e-12),
    ast.Pow: lambda a, b: np.power(np.clip(a, -1e12, 1e12), np.clip(b, -16, 16)),
    ast.Mod: lambda a, b: np.mod(a, b),
}
_ALLOWED_UNARYOPS: dict[type[ast.unaryop], Callable[[Any], Any]] = {ast.UAdd: lambda a: a, ast.USub: lambda a: -a}


class FormulaError(ValueError):
    pass


def evaluate_formula(expression: str, features: Mapping[str, np.ndarray], length: int | None = None) -> np.ndarray:
    """Evaluate a small, array-oriented expression without Python ``eval``."""

    text = str(expression).strip()
    if not text:
        raise FormulaError("Formula is empty.")
    target_length = int(length if length is not None else max((np.asarray(v).size for v in features.values()), default=1))
    names = {str(key): _align_vector(value, target_length) for key, value in features.items()}

    try:
        root = ast.parse(text, mode="eval")
    except SyntaxError as exc:
        raise FormulaError(f"Invalid formula syntax: {exc.msg}") from exc

    def visit(node: ast.AST) -> Any:
        if isinstance(node, ast.Expression):
            return visit(node.body)
        if isinstance(node, ast.Constant) and isinstance(node.value, int | float):
            return float(node.value)
        if isinstance(node, ast.Name):
            if node.id in names:
                return names[node.id]
            if node.id == "pi":
                return math.pi
            if node.id == "e":
                return math.e
            raise FormulaError(f"Unknown feature or constant: {node.id}")
        if isinstance(node, ast.BinOp) and type(node.op) in _ALLOWED_BINOPS:
            with np.errstate(all="ignore"):
                return _ALLOWED_BINOPS[type(node.op)](visit(node.left), visit(node.right))
        if isinstance(node, ast.UnaryOp) and type(node.op) in _ALLOWED_UNARYOPS:
            return _ALLOWED_UNARYOPS[type(node.op)](visit(node.operand))
        if isinstance(node, ast.Call) and isinstance(node.func, ast.Name):
            function = _FORMULA_FUNCTIONS.get(node.func.id)
            if function is None:
                raise FormulaError(f"Unsupported function: {node.func.id}")
            if node.keywords:
                raise FormulaError("Keyword arguments are not supported in formulas.")
            try:
                with np.errstate(all="ignore"):
                    return function(*(visit(argument) for argument in node.args))
            except Exception as exc:  # noqa: BLE001
                raise FormulaError(f"Could not evaluate {node.func.id}(): {exc}") from exc
        raise FormulaError(f"Unsupported formula element: {node.__class__.__name__}")

    value = visit(root)
    return _align_vector(value, target_length)


# ---------------------------------------------------------------------------
# Table analysis


def _read_delimited_table(path: Path) -> tuple[list[str], list[list[Any]]]:
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    sample = text[:8192]
    if path.suffix.lower() == ".tsv":
        dialect = csv.excel_tab
    else:
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        except csv.Error:
            dialect = csv.excel
    rows = list(csv.reader(io.StringIO(text), dialect))
    if not rows:
        raise ValueError("The table is empty.")
    first = rows[0]
    has_header = True
    with suppress(csv.Error):
        has_header = csv.Sniffer().has_header(sample)
    if has_header:
        headers = [str(value).strip() or f"Column {index + 1}" for index, value in enumerate(first)]
        data_rows = rows[1:]
    else:
        headers = [f"Input {index + 1}" for index in range(len(first))]
        data_rows = rows
    return headers, data_rows


def _read_xlsx_table(path: Path) -> tuple[list[str], list[list[Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency error path
        raise RuntimeError("XLSX import requires openpyxl.") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        first = next(iterator, None)
        if first is None:
            raise ValueError("The workbook's active sheet is empty.")
        headers = [str(value).strip() if value is not None else f"Column {index + 1}" for index, value in enumerate(first)]
        rows = [list(row) for row in iterator]
    finally:
        workbook.close()
    return headers, rows


def _coerce_numeric_columns(headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> tuple[dict[str, np.ndarray], dict[str, str]]:
    width = max(len(headers), max((len(row) for row in rows), default=0))
    if width <= 0:
        raise ValueError("The table does not contain columns.")
    padded_headers = list(headers) + [f"Column {index + 1}" for index in range(len(headers), width)]
    used: set[str] = set()
    result: dict[str, np.ndarray] = {}
    descriptions: dict[str, str] = {}
    for column_index in range(width):
        values: list[float] = []
        numeric = 0
        nonempty = 0
        for row in rows:
            raw = row[column_index] if column_index < len(row) else None
            if raw is None or str(raw).strip() == "":
                values.append(float("nan"))
                continue
            nonempty += 1
            try:
                values.append(float(raw))
                numeric += 1
            except (TypeError, ValueError):
                values.append(float("nan"))
        if not values or numeric < max(1, math.ceil(nonempty * 0.6)):
            continue
        array = np.asarray(values, dtype=np.float64)
        finite = np.isfinite(array)
        fill = float(np.nanmedian(array[finite])) if finite.any() else 0.0
        array = np.where(finite, array, fill).astype(np.float32)
        key = _safe_identifier(padded_headers[column_index], used)
        result[key] = array
        descriptions[key] = f'Imported numeric column "{padded_headers[column_index]}".'
    if not result:
        raise ValueError("No numeric columns were found. At least one mostly-numeric column is required.")
    return result, descriptions


def analysis_from_table_file(path: str | Path) -> AnalysisResult:
    source = Path(path).expanduser().resolve()
    if source.suffix.lower() == ".xlsx":
        headers, rows = _read_xlsx_table(source)
    else:
        headers, rows = _read_delimited_table(source)
    features, descriptions = _coerce_numeric_columns(headers, rows)
    row_count = max(array.size for array in features.values())
    features = {name: _align_vector(values, row_count) for name, values in features.items()}
    matrix = np.column_stack(list(features.values())).astype(np.float32)

    time_key = next((name for name in features if name in {"time", "timestamp", "seconds", "elapsed", "elapsed_seconds"}), None)
    if time_key is not None:
        times = features[time_key].astype(np.float32, copy=True)
        times -= float(times.min()) if times.size else 0.0
        if times.size > 1 and np.any(np.diff(times) < 0):
            times = np.arange(row_count, dtype=np.float32)
    else:
        times = np.arange(row_count, dtype=np.float32)
    duration = float(times[-1]) if times.size else 0.0

    magnitude = np.linalg.norm(matrix, axis=1).astype(np.float32)
    features["magnitude"] = magnitude
    features["column_mean"] = np.mean(matrix, axis=1).astype(np.float32)
    features["column_spread"] = np.std(matrix, axis=1).astype(np.float32)
    features["delta_magnitude"] = np.abs(np.diff(magnitude, prepend=magnitude[:1])).astype(np.float32)
    pcs = _principal_components(matrix, 3)
    features["pc1"], features["pc2"], features["pc3"] = pcs[:, 0], pcs[:, 1], pcs[:, 2]
    features["time"] = times
    features["frame"] = np.arange(row_count, dtype=np.float32)
    if duration > 0:
        features["time_normalized"] = times / duration
    else:
        features["time_normalized"] = np.zeros(row_count, dtype=np.float32)

    panel = _normalize_panel(matrix.T)
    chroma = np.zeros((12, row_count), dtype=np.float32)
    mfcc = np.zeros((20, row_count), dtype=np.float32)
    for index in range(12):
        chroma[index] = panel[index % panel.shape[0]] if panel.shape[0] else 0.0
    for index in range(20):
        mfcc[index] = panel[index % panel.shape[0]] if panel.shape[0] else 0.0

    # Keep every built-in audio-oriented mapping usable for tables. These are
    # explicitly compatibility aliases derived from imported numeric data, not
    # claims that a table contains acoustic measurements.
    energy = _normalize_panel(np.abs(magnitude))
    spread = _normalize_panel(features["column_spread"])
    change = _normalize_panel(features["delta_magnitude"])
    mean_signal = _normalize_panel(np.abs(features["column_mean"]))
    peak = max(float(np.max(energy)), 1e-6)
    features.setdefault("rms", energy.astype(np.float32))
    features.setdefault("rms_db", (20.0 * np.log10(np.clip(energy / peak, 1e-6, None))).astype(np.float32))
    signs = np.signbit(features["column_mean"])
    crossing = np.zeros(row_count, dtype=np.float32)
    if row_count > 1:
        crossing[1:] = (signs[1:] != signs[:-1]).astype(np.float32)
    features.setdefault("zero_crossing_rate", crossing)
    features.setdefault("spectral_centroid", (mean_signal * 20_000.0).astype(np.float32))
    features.setdefault("spectral_bandwidth", (spread * 10_000.0).astype(np.float32))
    features.setdefault("spectral_rolloff", (np.maximum(mean_signal, spread) * 22_000.0).astype(np.float32))
    features.setdefault("spectral_flatness", spread.astype(np.float32))
    features.setdefault("spectral_flux", change.astype(np.float32))
    features.setdefault("onset_strength", change.astype(np.float32))
    features.setdefault("dominant_frequency", (energy * 20_000.0).astype(np.float32))
    features.setdefault("chroma_mean", np.mean(chroma, axis=0).astype(np.float32))
    for index in range(chroma.shape[0]):
        features.setdefault(f"chroma_{index + 1}", chroma[index].astype(np.float32))
    for index in range(mfcc.shape[0]):
        features.setdefault(f"mfcc_{index + 1}", mfcc[index].astype(np.float32))

    descriptions.update(
        {key: FEATURE_DESCRIPTIONS.get(key, key.replace("_", " ").title()) for key in features if key not in descriptions}
    )

    return AnalysisResult(
        source_path=source,
        source_kind="table",
        times=times,
        duration=duration,
        features=features,
        sample_rate=0,
        audio_path=None,
        has_video=False,
        spectrogram=panel,
        spectrogram_frequencies=np.arange(panel.shape[0], dtype=np.float32),
        chromagram=chroma,
        mfcc=mfcc,
        waveform=features["column_mean"],
        metadata={"rows": row_count, "numeric_columns": list(features.keys())[: matrix.shape[1]], "original_headers": list(headers)},
        feature_descriptions=descriptions,
    )


# ---------------------------------------------------------------------------
# Media analysis


def _decode_audio(path: Path) -> tuple[np.ndarray, int]:
    """Decode mono floating-point audio with SoundFile, then FFmpeg fallback.

    SoundFile is fast for PCM/FLAC/Ogg sources and avoids a heavyweight feature
    stack during startup. FFmpeg provides broad container/codec coverage and is
    also the normal path for extracting audio from video.
    """

    try:
        import soundfile as sf

        samples, sample_rate = sf.read(str(path), dtype="float32", always_2d=False)
        array = np.asarray(samples, dtype=np.float32)
        if array.ndim > 1:
            array = np.mean(array, axis=1, dtype=np.float32)
        if array.size:
            return array.reshape(-1), int(sample_rate)
    except Exception:
        pass

    executable = shutil.which("ffmpeg")
    if not executable:
        raise RuntimeError("The source could not be decoded. Install FFmpeg for broad media support.")
    sample_rate = 44_100
    command = [
        executable,
        "-hide_banner",
        "-loglevel", "error",
        "-i", str(path),
        "-vn",
        "-ac", "1",
        "-ar", str(sample_rate),
        "-f", "f32le",
        "pipe:1",
    ]
    process = subprocess.run(command, capture_output=True, check=False)  # noqa: S603
    if process.returncode != 0 or not process.stdout:
        details = process.stderr.decode("utf-8", errors="replace")[-2000:]
        raise RuntimeError(f"The source has no decodable audio stream.\n{details}".strip())
    return np.frombuffer(process.stdout, dtype="<f4").astype(np.float32, copy=True), sample_rate


def _hz_to_mel(frequency: np.ndarray | float) -> np.ndarray:
    values = np.asarray(frequency, dtype=np.float64)
    return 2595.0 * np.log10(1.0 + np.maximum(values, 0.0) / 700.0)


def _mel_to_hz(mel: np.ndarray | float) -> np.ndarray:
    values = np.asarray(mel, dtype=np.float64)
    return 700.0 * (np.power(10.0, values / 2595.0) - 1.0)


def _mel_filterbank(
    sample_rate: int,
    n_fft: int,
    n_mels: int = 96,
    *,
    min_frequency: float = 0.0,
    max_frequency: float | None = None,
) -> tuple[np.ndarray, np.ndarray]:
    """Return a Slaney-normalized Mel bank for the requested frequency band."""

    frequencies = np.fft.rfftfreq(n_fft, d=1.0 / float(sample_rate))
    nyquist = sample_rate / 2.0
    minimum = float(np.clip(min_frequency, 0.0, max(0.0, nyquist - 1.0)))
    maximum = nyquist if max_frequency is None or max_frequency <= 0 else float(max_frequency)
    maximum = float(np.clip(maximum, minimum + 1.0, nyquist))
    mel_edges = np.linspace(_hz_to_mel(minimum), _hz_to_mel(maximum), n_mels + 2)
    hz_edges = _mel_to_hz(mel_edges)
    filters = np.zeros((n_mels, frequencies.size), dtype=np.float32)
    for index in range(n_mels):
        left, center, right = hz_edges[index : index + 3]
        if center > left:
            rising = (frequencies - left) / (center - left)
            filters[index] = np.maximum(filters[index], np.clip(rising, 0.0, 1.0))
        if right > center:
            falling = (right - frequencies) / (right - center)
            filters[index] = np.minimum(filters[index], np.clip(falling, 0.0, 1.0))
    widths = np.maximum(hz_edges[2 : n_mels + 2] - hz_edges[:n_mels], 1e-12)
    filters *= (2.0 / widths)[:, None].astype(np.float32)
    centers = hz_edges[1:-1].astype(np.float32)
    return filters, centers


def _resample_audio(samples: np.ndarray, source_rate: int, target_rate: int) -> np.ndarray:
    """Band-limited rational resampling without introducing a heavyweight DSP stack."""

    source_rate = int(source_rate)
    target_rate = int(target_rate)
    signal = np.asarray(samples, dtype=np.float32).reshape(-1)
    if source_rate <= 0 or target_rate <= 0 or source_rate == target_rate:
        return signal
    try:
        from scipy.signal import resample_poly
    except ImportError as exc:  # pragma: no cover - SciPy is a runtime requirement
        raise RuntimeError("Sample-rate conversion requires SciPy.") from exc
    divisor = gcd(source_rate, target_rate)
    up = target_rate // divisor
    down = source_rate // divisor
    return np.asarray(resample_poly(signal, up, down), dtype=np.float32).reshape(-1)


def _estimate_fundamental_frequency(
    magnitude: np.ndarray,
    frequencies: np.ndarray,
    *,
    min_frequency: float,
    max_frequency: float,
    harmonics: int = 5,
) -> np.ndarray:
    """Estimate frame fundamentals with bounded harmonic summation.

    The calculation is intentionally local and chunked. It provides a useful
    historical ``f0_hz`` mapping signal without adding the JIT-heavy analysis
    stack that made earlier builds unpredictable on ordinary computers.
    """

    spectrum = np.asarray(magnitude, dtype=np.float32)
    hz = np.asarray(frequencies, dtype=np.float64).reshape(-1)
    frame_count = spectrum.shape[1] if spectrum.ndim == 2 else 0
    output = np.zeros(frame_count, dtype=np.float32)
    if frame_count == 0 or hz.size < 2 or spectrum.shape[0] != hz.size:
        return output
    resolution = float(np.median(np.diff(hz))) if hz.size > 1 else 1.0
    minimum = max(float(min_frequency), max(20.0, resolution))
    maximum = min(float(max_frequency), float(hz[-1]))
    candidates = np.flatnonzero((hz >= minimum) & (hz <= maximum))
    if candidates.size == 0:
        return output
    candidate_hz = hz[candidates]
    harmonic_bins: list[tuple[np.ndarray, np.ndarray, float]] = []
    for harmonic in range(2, max(2, int(harmonics)) + 1):
        targets = candidate_hz * harmonic
        bins = np.searchsorted(hz, targets, side="left")
        valid = bins < hz.size
        if np.any(valid):
            harmonic_bins.append((np.flatnonzero(valid), bins[valid], 1.0 / math.sqrt(harmonic)))
    chunk_size = 256
    for offset in range(0, frame_count, chunk_size):
        stop = min(frame_count, offset + chunk_size)
        scores = spectrum[candidates, offset:stop].astype(np.float32, copy=True)
        for rows, bins, weight in harmonic_bins:
            scores[rows] += spectrum[bins, offset:stop] * np.float32(weight)
        best = np.argmax(scores, axis=0)
        selected = candidate_hz[best].astype(np.float32)
        peak = np.max(spectrum[:, offset:stop], axis=0)
        selected[peak <= 1e-10] = 0.0
        output[offset:stop] = selected
    return output


def _bounded_stft(
    samples: np.ndarray,
    sample_rate: int,
    *,
    n_fft: int = 2048,
    base_hop: int = 512,
    max_frames: int = 8192,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, int]:
    """Return magnitude, RMS, ZCR, times, and adaptive hop with bounded memory."""

    signal = np.asarray(samples, dtype=np.float32).reshape(-1)
    original_size = signal.size
    if signal.size < n_fft:
        signal = np.pad(signal, (0, n_fft - signal.size))
    available = max(0, signal.size - n_fft)
    hop = max(1, int(base_hop))
    expected = 1 + available // hop
    frequency_bins = n_fft // 2 + 1
    # Keep the primary STFT matrix around 128 MiB even when a creator selects
    # both a very large FFT and a very high frame ceiling.
    memory_limited_frames = max(128, 33_000_000 // max(1, frequency_bins))
    effective_max_frames = max(1, min(int(max_frames), memory_limited_frames))
    if expected > effective_max_frames and effective_max_frames > 1:
        hop = max(hop, int(math.ceil(available / (effective_max_frames - 1))))
    starts = np.arange(0, available + 1, hop, dtype=np.int64)
    if starts.size > effective_max_frames:
        starts = starts[:effective_max_frames]
    frame_count = max(1, int(starts.size))
    magnitude = np.empty((frequency_bins, frame_count), dtype=np.float32)
    rms = np.empty(frame_count, dtype=np.float32)
    zcr = np.empty(frame_count, dtype=np.float32)
    window = np.hanning(n_fft).astype(np.float32)
    sliding = np.lib.stride_tricks.sliding_window_view(signal, n_fft)
    chunk_size = 256
    for offset in range(0, frame_count, chunk_size):
        stop = min(frame_count, offset + chunk_size)
        batch = np.asarray(sliding[starts[offset:stop]], dtype=np.float32)
        rms[offset:stop] = np.sqrt(np.mean(np.square(batch, dtype=np.float32), axis=1) + 1e-20)
        signs = np.signbit(batch)
        zcr[offset:stop] = np.mean(signs[:, 1:] != signs[:, :-1], axis=1, dtype=np.float32)
        spectrum = np.fft.rfft(batch * window[None, :], axis=1)
        magnitude[:, offset:stop] = np.abs(spectrum).T.astype(np.float32)
    centers = starts.astype(np.float64) + n_fft * 0.5
    duration = original_size / float(sample_rate)
    times = np.minimum(centers / float(sample_rate), duration).astype(np.float32)
    return magnitude, rms, zcr, times, hop

def _video_only_analysis(path: Path, settings: AnalysisSettings | None = None) -> AnalysisResult:
    try:
        import cv2
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("Video analysis requires OpenCV.") from exc
    capture = cv2.VideoCapture(str(path))
    if not capture.isOpened():
        raise RuntimeError("The video could not be opened.")
    fps = float(capture.get(cv2.CAP_PROP_FPS) or 30.0)
    frame_total = int(capture.get(cv2.CAP_PROP_FRAME_COUNT) or 0)
    duration = frame_total / fps if frame_total > 0 and fps > 0 else 0.0
    requested = (settings or AnalysisSettings()).normalized()
    target_frames = min(requested.max_frames, max(1, frame_total)) if frame_total else requested.max_frames
    stride = max(1, int(math.ceil(frame_total / target_frames))) if frame_total else 1
    values: dict[str, list[float]] = {
        "brightness": [], "contrast": [], "saturation": [], "edge_density": [],
        "motion": [], "red_mean": [], "green_mean": [], "blue_mean": [],
    }
    times: list[float] = []
    previous: np.ndarray | None = None
    frame_index = 0
    try:
        while True:
            ok, frame = capture.read()
            if not ok:
                break
            if frame_index % stride:
                frame_index += 1
                continue
            reduced = cv2.resize(frame, (160, 90), interpolation=cv2.INTER_AREA)
            gray = cv2.cvtColor(reduced, cv2.COLOR_BGR2GRAY)
            hsv = cv2.cvtColor(reduced, cv2.COLOR_BGR2HSV)
            edges = cv2.Canny(gray, 80, 160)
            values["brightness"].append(float(np.mean(gray)))
            values["contrast"].append(float(np.std(gray)))
            values["saturation"].append(float(np.mean(hsv[:, :, 1])))
            values["edge_density"].append(float(np.mean(edges > 0)))
            values["motion"].append(float(np.mean(cv2.absdiff(gray, previous))) if previous is not None else 0.0)
            b, g, r = np.mean(reduced, axis=(0, 1))
            values["red_mean"].append(float(r))
            values["green_mean"].append(float(g))
            values["blue_mean"].append(float(b))
            times.append(frame_index / max(1e-9, fps))
            previous = gray
            frame_index += 1
            if len(times) >= target_frames:
                break
    finally:
        capture.release()
    if not times:
        raise RuntimeError("No video frames could be decoded.")
    time_array = np.asarray(times, dtype=np.float32)
    feature_map = {name: np.asarray(series, dtype=np.float32) for name, series in values.items()}
    matrix = np.column_stack(list(feature_map.values()))
    pcs = _principal_components(matrix, 3)
    feature_map.update({"pc1": pcs[:, 0], "pc2": pcs[:, 1], "pc3": pcs[:, 2]})
    feature_map["time"] = time_array
    feature_map["frame"] = np.arange(time_array.size, dtype=np.float32)
    feature_map["time_normalized"] = time_array / max(duration, float(time_array[-1]), 1e-12)
    feature_map["magnitude"] = np.linalg.norm(matrix, axis=1).astype(np.float32)
    feature_map["column_mean"] = np.mean(matrix, axis=1).astype(np.float32)
    feature_map["column_spread"] = np.std(matrix, axis=1).astype(np.float32)
    feature_map["delta_magnitude"] = np.abs(
        np.diff(feature_map["magnitude"], prepend=feature_map["magnitude"][:1])
    ).astype(np.float32)
    # Aliases keep audio-oriented presets useful on silent video.
    feature_map["rms"] = _normalize_panel(feature_map["brightness"])
    feature_map["rms_db"] = 20.0 * np.log10(np.clip(feature_map["rms"], 1e-6, None))
    feature_map["spectral_flux"] = feature_map["motion"]
    feature_map["onset_strength"] = feature_map["motion"]
    feature_map["spectral_centroid"] = feature_map["brightness"]
    feature_map["spectral_bandwidth"] = feature_map["contrast"]
    feature_map["spectral_rolloff"] = feature_map["edge_density"] * 22_000.0
    feature_map["spectral_flatness"] = _normalize_panel(feature_map["contrast"])
    feature_map["zero_crossing_rate"] = feature_map["edge_density"]
    feature_map["dominant_frequency"] = feature_map["edge_density"] * 20_000.0
    feature_map["chroma_mean"] = feature_map["saturation"]
    panel = _normalize_panel(matrix.T)
    chroma = np.vstack([panel[index % panel.shape[0]] for index in range(12)]).astype(np.float32)
    mfcc = np.vstack([panel[index % panel.shape[0]] for index in range(requested.mfcc_count)]).astype(np.float32)
    for index in range(chroma.shape[0]):
        feature_map[f"chroma_{index + 1}"] = chroma[index]
    for index in range(mfcc.shape[0]):
        feature_map[f"mfcc_{index + 1}"] = mfcc[index]
    descriptions = {name: f"Video-derived {name.replace('_', ' ')}." for name in feature_map}
    descriptions.update(FEATURE_DESCRIPTIONS)
    return AnalysisResult(
        source_path=path,
        source_kind="video",
        times=time_array,
        duration=max(duration, float(time_array[-1])),
        features=feature_map,
        sample_rate=0,
        audio_path=None,
        has_video=True,
        spectrogram=panel,
        spectrogram_frequencies=np.arange(panel.shape[0], dtype=np.float32),
        chromagram=chroma,
        mfcc=mfcc,
        waveform=feature_map["brightness"],
        metadata={
            "video_fps": fps,
            "frame_count": frame_total,
            "analysis_mode": "video-only",
            "analysis_settings": requested.to_dict(),
            "analysis_engine": ANALYSIS_ENGINE_VERSION,
        },
        feature_descriptions=descriptions,
    )


def analyze_media(
    path: str | Path,
    settings: AnalysisSettings | Mapping[str, Any] | None = None,
) -> AnalysisResult:
    """Analyze audio or a video's audio track with creator-controlled local DSP.

    FFT size, hop, sample rate, frequency band, Mel count, MFCC count, and the
    frame ceiling are all functional settings. Long sources still receive an
    adaptive hop so analysis memory remains bounded on ordinary computers.
    """

    source = Path(path).expanduser().resolve()
    if not source.is_file():
        raise FileNotFoundError(source)
    has_video = source.suffix.lower() in VIDEO_EXTENSIONS
    requested = settings if isinstance(settings, AnalysisSettings) else AnalysisSettings.from_mapping(settings)
    try:
        samples, source_sample_rate = _decode_audio(source)
    except RuntimeError:
        if has_video:
            return _video_only_analysis(source, requested)
        raise

    samples = np.asarray(samples, dtype=np.float32).reshape(-1)
    if samples.size < 16:
        raise ValueError("The source is too short to analyze.")
    source_sample_count = int(samples.size)
    source_duration = float(source_sample_count / source_sample_rate)
    configured = requested.normalized(native_sample_rate=source_sample_rate)
    sample_rate = int(configured.sample_rate or source_sample_rate)
    if sample_rate != source_sample_rate:
        samples = _resample_audio(samples, source_sample_rate, sample_rate)
    if samples.size < 16:
        raise ValueError("The resampled source is too short to analyze.")
    peak = float(np.max(np.abs(samples)))
    if peak > 1.0:
        samples = samples / peak

    n_fft = configured.n_fft
    magnitude, rms, zcr, times, hop = _bounded_stft(
        samples,
        sample_rate,
        n_fft=n_fft,
        base_hop=configured.hop_length,
        max_frames=configured.max_frames,
    )
    frame_count = magnitude.shape[1]
    full_frequencies = np.fft.rfftfreq(n_fft, d=1.0 / float(sample_rate)).astype(np.float32)
    nyquist = sample_rate / 2.0
    minimum_frequency = float(np.clip(configured.min_frequency, 0.0, max(0.0, nyquist - 1.0)))
    maximum_frequency = nyquist if configured.max_frequency <= 0 else float(configured.max_frequency)
    maximum_frequency = float(np.clip(maximum_frequency, minimum_frequency + 1.0, nyquist))
    band_mask = (full_frequencies >= minimum_frequency) & (full_frequencies <= maximum_frequency)
    if not np.any(band_mask):
        nearest = int(np.argmin(np.abs(full_frequencies - minimum_frequency)))
        band_mask[nearest] = True
    frequencies = full_frequencies[band_mask]
    magnitude = np.ascontiguousarray(magnitude[band_mask], dtype=np.float32)
    energy = np.maximum(np.sum(magnitude, axis=0), 1e-20)

    centroid = np.sum(magnitude * frequencies[:, None], axis=0) / energy
    second_moment = np.sum(magnitude * np.square(frequencies)[:, None], axis=0) / energy
    bandwidth = np.sqrt(np.maximum(second_moment - np.square(centroid), 0.0))
    dominant = frequencies[np.argmax(magnitude, axis=0)]
    fundamental = _estimate_fundamental_frequency(
        magnitude,
        frequencies,
        min_frequency=minimum_frequency,
        max_frequency=maximum_frequency,
    )

    normalized_spectrum = magnitude / np.maximum(np.linalg.norm(magnitude, axis=0, keepdims=True), 1e-20)
    spectral_delta = np.diff(normalized_spectrum, axis=1, prepend=normalized_spectrum[:, :1])
    flux = np.sqrt(np.sum(np.square(spectral_delta, dtype=np.float32), axis=0))
    onset = np.sum(np.maximum(spectral_delta, 0.0), axis=0)
    del normalized_spectrum, spectral_delta

    np.square(magnitude, out=magnitude)
    power = magnitude
    cumulative = np.cumsum(power, axis=0)
    total_power = np.maximum(cumulative[-1], 1e-20)
    rolloff_indices = np.argmax(cumulative >= (0.85 * total_power)[None, :], axis=0)
    rolloff = frequencies[rolloff_indices]
    del cumulative
    flatness = np.exp(np.mean(np.log(power + 1e-20), axis=0)) / np.maximum(np.mean(power, axis=0), 1e-20)

    chroma = np.zeros((12, frame_count), dtype=np.float32)
    valid_bins = np.flatnonzero(frequencies > 0.0)
    if valid_bins.size:
        midi = np.rint(69.0 + 12.0 * np.log2(frequencies[valid_bins] / 440.0)).astype(np.int32)
        pitch_classes = np.mod(midi, 12)
        for pitch_class in range(12):
            bins = valid_bins[pitch_classes == pitch_class]
            if bins.size:
                chroma[pitch_class] = np.sum(power[bins], axis=0)
    chroma /= np.maximum(np.sum(chroma, axis=0, keepdims=True), 1e-20)

    mel_filters, mel_centers = _mel_filterbank(
        sample_rate,
        n_fft,
        configured.n_mels,
        min_frequency=minimum_frequency,
        max_frequency=maximum_frequency,
    )
    mel_filters = np.ascontiguousarray(mel_filters[:, band_mask], dtype=np.float32)
    mel_power = np.empty((mel_filters.shape[0], frame_count), dtype=np.float32)
    for offset in range(0, frame_count, 256):
        stop = min(frame_count, offset + 256)
        mel_power[:, offset:stop] = mel_filters @ power[:, offset:stop]
    np.maximum(mel_power, 1e-20, out=mel_power)
    del power, mel_filters
    mel_reference = max(float(np.max(mel_power)), 1e-20)
    mel_db = (10.0 * np.log10(mel_power / mel_reference)).astype(np.float32)
    try:
        from scipy.fft import dct
    except ImportError as exc:  # pragma: no cover - SciPy is a runtime requirement
        raise RuntimeError("Media analysis requires SciPy.") from exc
    mfcc = dct(mel_db, type=2, axis=0, norm="ortho")[: configured.mfcc_count].astype(np.float32)
    rms_db = (20.0 * np.log10(np.maximum(rms, 1e-12) / max(float(np.max(rms)), 1e-12))).astype(np.float32)
    duration = max(source_duration, float(samples.size / sample_rate))

    features: dict[str, np.ndarray] = {
        "time": times,
        "frame": np.arange(frame_count, dtype=np.float32),
        "time_normalized": times / max(duration, 1e-12),
        "rms": _align_vector(rms, frame_count),
        "rms_db": _align_vector(rms_db, frame_count),
        "zero_crossing_rate": _align_vector(zcr, frame_count),
        "spectral_centroid": _align_vector(centroid, frame_count),
        "spectral_bandwidth": _align_vector(bandwidth, frame_count),
        "spectral_rolloff": _align_vector(rolloff, frame_count),
        "spectral_flatness": _align_vector(flatness, frame_count),
        "spectral_flux": _align_vector(flux, frame_count),
        "onset_strength": _align_vector(onset, frame_count),
        "dominant_frequency": _align_vector(dominant, frame_count),
        "fundamental_frequency": _align_vector(fundamental, frame_count),
        "chroma_mean": _align_vector(np.mean(chroma, axis=0), frame_count),
    }
    for index in range(chroma.shape[0]):
        features[f"chroma_{index + 1}"] = _align_vector(chroma[index], frame_count)
    for index in range(mfcc.shape[0]):
        features[f"mfcc_{index + 1}"] = _align_vector(mfcc[index], frame_count)

    pca_columns = [
        features["rms"],
        features["zero_crossing_rate"],
        features["spectral_centroid"],
        features["spectral_bandwidth"],
        features["spectral_rolloff"],
        features["spectral_flatness"],
        features["spectral_flux"],
        features["onset_strength"],
        features["chroma_mean"],
    ]
    pca_columns.extend(features.get(f"mfcc_{index}", np.zeros(frame_count, dtype=np.float32)) for index in range(1, 8))
    pca_inputs = np.column_stack(pca_columns)
    magnitude_feature = np.linalg.norm(pca_inputs, axis=1).astype(np.float32)
    features["magnitude"] = magnitude_feature
    features["column_mean"] = np.mean(pca_inputs, axis=1).astype(np.float32)
    features["column_spread"] = np.std(pca_inputs, axis=1).astype(np.float32)
    features["delta_magnitude"] = np.abs(
        np.diff(magnitude_feature, prepend=magnitude_feature[:1])
    ).astype(np.float32)
    pcs = _principal_components(pca_inputs, 3)
    features["pc1"], features["pc2"], features["pc3"] = pcs[:, 0], pcs[:, 1], pcs[:, 2]

    descriptions = dict(FEATURE_DESCRIPTIONS)
    descriptions.update(
        {f"mfcc_{index + 1}": f"Mel-frequency cepstral coefficient {index + 1}." for index in range(configured.mfcc_count)}
    )
    descriptions.update({f"chroma_{index + 1}": f"Pitch-class energy channel {index + 1}." for index in range(12)})

    max_waveform = 240_000
    if samples.size > max_waveform:
        indices = np.linspace(0, samples.size - 1, max_waveform, dtype=np.int64)
        waveform = samples[indices]
    else:
        waveform = samples

    return AnalysisResult(
        source_path=source,
        source_kind="video" if has_video else "audio",
        times=times,
        duration=duration,
        features=features,
        sample_rate=sample_rate,
        audio_path=source,
        has_video=has_video,
        spectrogram=_normalize_panel(mel_db),
        spectrogram_frequencies=mel_centers,
        chromagram=_normalize_panel(chroma),
        mfcc=_normalize_panel(mfcc),
        waveform=waveform.astype(np.float32),
        metadata={
            "sample_rate": sample_rate,
            "source_sample_rate": int(source_sample_rate),
            "samples": int(samples.size),
            "source_samples": source_sample_count,
            "analysis_frames": frame_count,
            "max_analysis_frames": configured.max_frames,
            "requested_hop_length": configured.hop_length,
            "hop_length": hop,
            "n_fft": n_fft,
            "min_frequency": minimum_frequency,
            "max_frequency": maximum_frequency,
            "n_mels": configured.n_mels,
            "mfcc_count": configured.mfcc_count,
            "analysis_settings": configured.to_dict(),
            "analysis_engine": ANALYSIS_ENGINE_VERSION,
        },
        feature_descriptions=descriptions,
    )


# ---------------------------------------------------------------------------
# Geometry construction


def _normalize(values: np.ndarray, mode: str) -> np.ndarray:
    array = _vector(values).astype(np.float64)
    if array.size == 0:
        return array.astype(np.float32)
    if mode == "raw":
        return np.clip(array, -1e6, 1e6).astype(np.float32)
    if mode == "minmax":
        low, high = np.percentile(array, [1.0, 99.0])
        if high <= low + 1e-12:
            return np.zeros(array.size, dtype=np.float32)
        return (np.clip((array - low) / (high - low), 0.0, 1.0) * 2.0 - 1.0).astype(np.float32)
    median = float(np.median(array))
    q25, q75 = np.percentile(array, [25.0, 75.0])
    scale = float((q75 - q25) / 1.349)
    if scale <= 1e-12:
        scale = float(np.std(array))
    if scale <= 1e-12:
        return np.zeros(array.size, dtype=np.float32)
    return np.clip((array - median) / scale, -6.0, 6.0).astype(np.float32)


def _color_map(values: np.ndarray, name: str) -> np.ndarray:
    array = _vector(values)
    if array.size == 0:
        return np.empty((0, 4), dtype=np.float32)
    lo, hi = np.percentile(array, [1.0, 99.0])
    normalized = np.zeros_like(array) if hi <= lo + 1e-12 else np.clip((array - lo) / (hi - lo), 0.0, 1.0)
    try:
        from matplotlib import colormaps

        cmap = colormaps.get_cmap(name)
        rgba = np.asarray(cmap(normalized), dtype=np.float32)
    except Exception:
        rgba = np.column_stack((normalized, 0.3 + 0.7 * (1.0 - normalized), 0.8 * np.ones_like(normalized), np.ones_like(normalized)))
    return rgba.astype(np.float32)


def build_geometry(
    analysis: AnalysisResult,
    x_formula: str,
    y_formula: str,
    z_formula: str,
    color_formula: str,
    size_formula: str,
    *,
    normalize_mode: str = "zscore",
    max_points: int = 3000,
    low_volume_cutoff_db: float = 0.0,
    colormap: str = "plasma",
) -> GeometryResult:
    length = int(analysis.times.size)
    if length <= 0:
        raise ValueError("The analysis contains no frames.")
    formulas = {
        "x": str(x_formula), "y": str(y_formula), "z": str(z_formula),
        "color": str(color_formula), "size": str(size_formula),
    }
    raw = {key: evaluate_formula(expression, analysis.features, length) for key, expression in formulas.items()}
    valid = np.ones(length, dtype=bool)
    for values in raw.values():
        valid &= np.isfinite(values)
    if low_volume_cutoff_db > 0:
        if "rms_db" in analysis.features:
            level_db = _align_vector(analysis.features["rms_db"], length)
        elif "rms" in analysis.features:
            rms = _align_vector(analysis.features["rms"], length)
            level_db = 20.0 * np.log10(np.clip(rms, 1e-12, None) / max(float(np.max(rms)), 1e-12))
        else:
            level_db = np.zeros(length, dtype=np.float32)
        valid &= level_db >= -float(low_volume_cutoff_db)
    source_indices = np.flatnonzero(valid)
    if source_indices.size == 0:
        raise ValueError("No frames remain after filtering. Reduce the low-volume cutoff or change the formulas.")

    x_full = _normalize(raw["x"][valid], normalize_mode)
    y_full = _normalize(raw["y"][valid], normalize_mode)
    z_full = _normalize(raw["z"][valid], normalize_mode)
    color_full = raw["color"][valid].astype(np.float32)
    size_raw = np.abs(raw["size"][valid].astype(np.float64))
    size_lo, size_hi = np.percentile(size_raw, [5.0, 95.0])
    if size_hi <= size_lo + 1e-12:
        size_full = np.ones(size_raw.size, dtype=np.float32)
    else:
        size_full = (0.25 + 1.75 * np.clip((size_raw - size_lo) / (size_hi - size_lo), 0.0, 1.0)).astype(np.float32)
    times_full = analysis.times[valid].astype(np.float32)
    rgba_full = _color_map(color_full, colormap)

    target = max(1, min(int(max_points), source_indices.size))
    if target < source_indices.size:
        plot_positions = np.unique(np.linspace(0, source_indices.size - 1, target, dtype=np.int64))
    else:
        plot_positions = np.arange(source_indices.size, dtype=np.int64)
    return GeometryResult(
        x_full=x_full,
        y_full=y_full,
        z_full=z_full,
        color_full=color_full,
        size_full=size_full,
        times_full=times_full,
        rgba_full=rgba_full,
        source_indices_full=source_indices,
        x_plot=x_full[plot_positions],
        y_plot=y_full[plot_positions],
        z_plot=z_full[plot_positions],
        color_plot=color_full[plot_positions],
        size_plot=size_full[plot_positions],
        times_plot=times_full[plot_positions],
        rgba_plot=rgba_full[plot_positions],
        source_indices_plot=source_indices[plot_positions],
        formulas=formulas,
        normalize_mode=str(normalize_mode),
        colormap=str(colormap),
    )


def format_feature_reference(analysis: AnalysisResult) -> str:
    lines = [
        f"Source: {analysis.source_path}",
        f"Kind: {analysis.source_kind}",
        f"Duration: {analysis.duration:.3f} seconds",
        f"Analysis frames: {analysis.times.size:,}",
    ]
    if analysis.sample_rate:
        lines.append(f"Sample rate: {analysis.sample_rate:,} Hz")
    lines.extend(["", "Formula-ready features", "----------------------"])
    for name in sorted(analysis.features):
        values = analysis.features[name]
        description = analysis.feature_descriptions.get(name) or FEATURE_DESCRIPTIONS.get(name) or name.replace("_", " ").title()
        if values.size:
            lines.append(f"{name:<28} {description}  [min {float(np.min(values)):.4g}, max {float(np.max(values)):.4g}]")
        else:
            lines.append(f"{name:<28} {description}")
    lines.extend(
        [
            "",
            "Operators: +  -  *  /  **  %",
            "Functions: abs, sqrt, log, log1p, exp, clip, smooth, mean, avg, sum, max, min",
            "Examples: smooth(spectral_flux, 5) · 0.7*mfcc_1 + 0.3*chroma_mean",
        ]
    )
    return "\n".join(lines)


__all__ = [
    "AnalysisResult",
    "AnalysisSettings",
    "AUDIO_EXTENSIONS",
    "DEFAULT_PRESETS",
    "FEATURE_DESCRIPTIONS",
    "FormulaError",
    "GeometryResult",
    "TABLE_EXTENSIONS",
    "VIDEO_EXTENSIONS",
    "analysis_from_table_file",
    "analyze_media",
    "build_geometry",
    "evaluate_formula",
    "format_feature_reference",
    "is_table_file",
]
